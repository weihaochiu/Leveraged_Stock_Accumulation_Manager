from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import load_workbook

from stock_manager.database import Database, PortfolioRepository
from stock_manager.import_export import BackupService, ExcelService


class BackupExportTests(unittest.TestCase):
    def test_complete_excel_and_backup_package(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp); db = Database(root / "data" / "portfolio.db"); repo = PortfolioRepository(db)
            security = repo.ensure_security("2330", "台積電")
            repo.add_buy_lot({"security_id": security, "buy_date": "2026-01-01", "buy_price": 1000, "original_shares": 10,
                "stock_amount": 10000, "buy_fee": 20, "funding_type": "CASH", "cash_funded": 10020, "loan_funded": 0,
                "target_return_pct": 15, "recovery_tolerance_amount": 100})
            output = ExcelService(repo).export_complete(root / "export.xlsx")
            wb = load_workbook(output, read_only=True)
            self.assertIn("交易總覽", wb.sheetnames); self.assertIn("買入批次", wb.sheetnames); wb.close()
            service = BackupService(db, repo, root / "stage"); service.ensure_default_target(root / "backups")
            result = service.run("TEST")
            self.assertEqual(result["status"], "SUCCESS")
            destination = Path(result["results"][0]["path"])
            self.assertTrue((destination / "portfolio.db").exists())
            self.assertTrue((destination / "portfolio.xlsx").exists())
            self.assertTrue((destination / "backup_manifest.json").exists())


if __name__ == "__main__":
    unittest.main()
