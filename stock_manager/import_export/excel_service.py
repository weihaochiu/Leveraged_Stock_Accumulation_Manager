from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from stock_manager.database import PortfolioRepository
from stock_manager.domain import zh


SHEETS = [
    ("買入批次", "buy_lots"), ("賣出紀錄", "sell_transactions"),
    ("股利紀錄", "dividends"), ("公司行動", "corporate_actions"),
    ("貸款", "loans"), ("貸款交易", "loan_transactions"),
    ("股票資料", "securities"), ("股票策略", "security_strategies"),
    ("股價紀錄", "price_history"), ("券商帳戶", "broker_accounts"),
    ("對帳紀錄", "reconciliations"), ("OCR匯入紀錄", "ocr_drafts"),
    ("設定摘要", "settings"), ("稽核紀錄", "audit_log"),
]

MASTER_HEADERS = {
    "id": "買進批次", "buy_date": "買入日期", "broker_name": "券商", "symbol": "股票代號",
    "security_name": "股票名稱", "funding_type": "資金來源", "loan_id": "貸款編號",
    "buy_price": "買入價格", "original_shares": "原始股數", "original_capital": "原始總投入",
    "current_price": "現價", "current_return_pct": "目前報酬率%", "target_return_pct": "目標報酬率%",
    "target_price": "目標價格", "distance_to_target_pct": "距離目標%", "strategy_status": "策略狀態",
    "net_cash_recovered": "累積淨回收", "capital_recovery_ratio": "本金回收率", "full_cost_recovery_ratio": "完整成本回收率",
    "remaining_capital_at_risk": "尚未回收本金", "recovery_difference": "回本差額", "remaining_shares": "目前股數",
    "free_shares": "已回本持股", "market_value": "目前市值", "free_share_value": "已回本持股市值",
    "loan_interest": "累積利息", "last_sell_date": "最近賣出日期", "holding_days": "持有天數", "updated_at": "最後修改時間",
}

RAW_HEADER_ZH = {
    "id": "編號", "security_id": "股票資料編號", "broker_account_id": "券商帳戶編號", "lot_id": "買進批次編號",
    "loan_id": "貸款編號", "sell_id": "賣出編號", "symbol": "股票代號", "name": "名稱", "market": "市場", "currency": "幣別",
    "buy_date": "買入日期", "sell_date": "賣出日期", "buy_price": "買入價格", "sell_price": "賣出價格", "price": "價格",
    "original_shares": "原始股數", "shares": "股數", "stock_amount": "股票成交金額", "gross_amount": "成交總額",
    "buy_fee": "買入手續費", "commission": "手續費", "tax": "交易稅／稅額", "other_fee": "其他費用", "other_cost": "其他買入成本",
    "net_cash": "實際淨回收", "net_amount": "實收金額", "original_capital": "原始總投入", "funding_type": "資金來源",
    "loan_funded": "貸款投入", "cash_funded": "自有資金投入", "target_return_pct": "目標報酬率%", "recovery_mode": "回收模式",
    "recovery_tolerance_amount": "回本容許差額", "recovery_tolerance_pct": "回本容許比例%", "broker_order_id": "券商下單編號",
    "broker_execution_id": "券商成交編號", "borrow_date": "借款日期", "original_principal": "原始本金", "annual_interest_rate": "年利率",
    "transaction_date": "交易日期", "transaction_type": "交易類型", "amount": "金額", "payment_date": "入帳日", "ex_date": "除息／除權日",
    "base_shares": "基準股數", "dividend_per_share": "每股股利", "dividend_type": "股利類型", "include_in_recovery": "是否計入回本",
    "action_type": "公司行動類型", "effective_date": "生效日期", "share_change": "股數增減", "cash_amount": "現金金額", "ratio": "比例",
    "price_date": "價格日期", "source": "來源", "updated_at": "最後修改時間", "created_at": "建立時間", "note": "備註", "active": "啟用",
    "key": "設定項目", "value": "設定值", "occurred_at": "時間", "action": "動作", "entity_type": "資料類型", "entity_id": "資料編號",
    "old_value": "修改前", "new_value": "修改後", "status": "狀態", "session_date": "對帳日期", "system_shares": "系統持股",
    "broker_shares": "券商持股", "difference": "差異", "path": "路徑", "enabled": "啟用", "target_type": "位置類型",
}


class ExcelService:
    def __init__(self, repository: PortfolioRepository):
        self.repository = repository

    def export_complete(self, path: str | Path) -> Path:
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        wb.remove(wb.active)
        dashboard = self.repository.dashboard()
        ws = wb.create_sheet("策略總覽")
        ws.append(["項目", "數值"])
        labels = {
            "market_value": "股票總市值", "original_capital": "原始累積投入", "net_cash_recovered": "累積回收現金",
            "capital_at_risk": "尚未回收本金", "free_share_value": "已回本持股市值", "cash_surplus": "超額回收現金",
            "dividends": "累積現金股利", "loan_interest": "累積貸款利息", "loan_balance": "目前貸款餘額",
            "recovered_lots": "已回本批次數", "total_lots": "全部批次數",
        }
        for key, label in labels.items():
            ws.append([label, float(dashboard[key]) if hasattr(dashboard[key], "as_tuple") else dashboard[key]])
        self._style(ws)

        ws = wb.create_sheet("交易總覽")
        rows = self.repository.master_rows()
        keys = list(MASTER_HEADERS)
        ws.append([MASTER_HEADERS[k] for k in keys])
        for row in rows:
            values = []
            for key in keys:
                value = row.get(key)
                if hasattr(value, "value"):
                    value = zh(value)
                elif hasattr(value, "as_tuple"):
                    value = float(value)
                values.append(value)
            ws.append(values)
        self._style(ws)

        for title, table in SHEETS:
            ws = wb.create_sheet(title)
            data = self.repository.table_rows(table)
            if data:
                columns = list(data[0])
                ws.append([RAW_HEADER_ZH.get(column, column) for column in columns])
                for row in data:
                    ws.append([row.get(c) for c in columns])
            else:
                ws.append(["尚無資料"])
            self._style(ws)
        wb.save(path)
        load_workbook(path, read_only=True).close()
        return path

    def export_master_csv(self, path: str | Path, rows: Iterable[dict] | None = None) -> Path:
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        keys = list(MASTER_HEADERS)
        with path.open("w", newline="", encoding="utf-8-sig") as fh:
            writer = csv.writer(fh)
            writer.writerow([MASTER_HEADERS[k] for k in keys])
            for row in rows or self.repository.master_rows():
                writer.writerow([zh(row.get(k)) if hasattr(row.get(k), "value") else row.get(k) for k in keys])
        return path

    def preview(self, path: str | Path, sheet_name: str | None = None, limit: int = 100) -> tuple[list[str], list[list]]:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True, max_row=limit + 1))
        wb.close()
        if not rows:
            return [], []
        return [str(x or "") for x in rows[0]], [list(r) for r in rows[1:]]

    @staticmethod
    def _style(ws) -> None:
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2F5597")
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for column in range(1, min(ws.max_column, 40) + 1):
            lengths = [len(str(ws.cell(r, column).value or "")) for r in range(1, min(ws.max_row, 200) + 1)]
            ws.column_dimensions[get_column_letter(column)].width = min(35, max(10, max(lengths, default=10) + 2))
